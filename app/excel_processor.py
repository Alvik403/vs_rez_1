from __future__ import annotations

import io
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INPUT_SHEET_NAME = "Вскрытие резервов"
TARGET_SHEET_NAME = "Справка по резервам "
FLAT_SHEET_NAME = "Справка"
SUMMARY_SHEET_NAME = "СВОД общий"
RESERVE_LEVELS = ("РП", "РГП", "ЗГД", "ГД")


@dataclass
class ProjectSummary:
    project: str
    plan: float = 0
    description: str = ""
    approved_reserves: dict[str, float] = field(default_factory=lambda: {level: 0 for level in RESERVE_LEVELS})
    openings: list["OpeningRow"] = field(default_factory=list)


@dataclass
class OpeningRow:
    source_file: str
    project: str
    cfo: str
    level: str
    date_year: int | str
    reserve_by_pd: float | str
    amount: float
    reason: str
    counterparty: str
    opening_type: str
    contract: str
    remainder: float | str
    economist_comment: str

    @property
    def work_type(self) -> str:
        return self.economist_comment or self.reason or self.opening_type

    @property
    def comment(self) -> str:
        parts = [
            f"Причина: {self.reason}" if self.reason else "",
            f"Контрагент: {self.counterparty}" if self.counterparty else "",
            f"Договор: {self.contract}" if self.contract else "",
            f"Остаток от резерва: {self.remainder}" if self.remainder not in ("", None) else "",
        ]
        return "; ".join(part for part in parts if part)


@dataclass
class ParsedFile:
    source_file: str
    projects: dict[str, ProjectSummary]
    openings: list[OpeningRow]


def normalize(value: Any) -> str:
    return (
        str(value or "")
        .replace("\xa0", " ")
        .replace("ё", "е")
        .lower()
        .strip()
    )


def normalized_cell(cell: Cell | None) -> str:
    return normalize(cell.value if cell else "")


def text_value(cell: Cell | None) -> str:
    value = cell.value if cell else ""
    if value is None:
        return ""
    return str(value).strip()


def numeric_value(cell: Cell | None) -> float | None:
    if not cell or cell.value in ("", None):
        return None
    value = cell.value
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.startswith("="):
        return eval_simple_formula(value)

    raw = str(value).replace("\xa0", "").replace(" ", "").strip()
    if not raw:
        return None
    decimal_like = bool(re.fullmatch(r"-?\d+,\d{1,2}", raw))
    normalized = raw.replace(",", ".") if decimal_like else raw.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def eval_simple_formula(formula: str) -> float | None:
    expression = formula.lstrip("=").replace(" ", "").replace(",", ".")
    if not re.fullmatch(r"[\d.+\-*/()]+", expression):
        return None
    try:
        result = eval(expression, {"__builtins__": {}}, {})
    except Exception:
        return None
    return float(result) if isinstance(result, (int, float)) else None


def date_year(value: Any) -> int | str:
    if value in ("", None):
        return ""
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, (int, float)):
        # Excel serial date, 25569 days from unix epoch.
        try:
            return datetime.fromtimestamp((float(value) - 25569) * 86400).year
        except Exception:
            return ""
    year_match = re.search(r"\b(19\d{2}|20\d{2})\b", str(value))
    if year_match:
        return int(year_match.group(1))
    for fmt in ("%d.%m.%Y", "%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt).year
        except ValueError:
            pass
    return ""


def find_sheet(workbook: Workbook, wanted_name: str) -> Worksheet | None:
    wanted = normalize(wanted_name)
    for name in workbook.sheetnames:
        if normalize(name) == wanted:
            return workbook[name]
    for name in workbook.sheetnames:
        if wanted in normalize(name):
            return workbook[name]
    return None


def row_labels(sheet: Worksheet, row: int) -> dict[int, str]:
    return {col: normalized_cell(sheet.cell(row=row, column=col)) for col in range(1, sheet.max_column + 1)}


def find_row(sheet: Worksheet, predicate) -> int:
    for row in range(1, sheet.max_row + 1):
        labels = row_labels(sheet, row)
        if predicate(labels, row):
            return row
    return -1


def find_column(labels: dict[int, str], label: str, start_col: int = 1) -> int:
    needle = normalize(label)
    for col, value in sorted(labels.items()):
        if col < start_col:
            continue
        v = normalize(value)
        if v == needle:
            return col
    for col, value in sorted(labels.items()):
        if col < start_col:
            continue
        v = normalize(value)
        if not v:
            continue
        if needle in v:
            if v != needle and needle == "гд" and v == "згд":
                continue
            return col
    return -1


def safe_cell(sheet: Worksheet, row: int, column: int) -> Cell | None:
    return sheet.cell(row=row, column=column) if column > 0 else None


def apply_formula_fallbacks(value_sheet: Worksheet, formula_sheet: Worksheet | None) -> None:
    """Если Excel не сохранил cached value, пробуем посчитать простые формулы сами."""
    if formula_sheet is None:
        return
    max_row = min(value_sheet.max_row, formula_sheet.max_row)
    max_col = min(value_sheet.max_column, formula_sheet.max_column)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value_cell = value_sheet.cell(row=row, column=col)
            if value_cell.value not in ("", None):
                continue
            formula_value = formula_sheet.cell(row=row, column=col).value
            if isinstance(formula_value, str) and formula_value.startswith("="):
                fallback = eval_simple_formula(formula_value)
                if fallback is not None:
                    value_cell.value = fallback


def read_project_summaries(sheet: Worksheet, detail_header_row: int) -> dict[str, ProjectSummary]:
    summary_header_row = find_row(
        sheet,
        lambda labels, row: row < detail_header_row and any("наименование проекта" in value for value in labels.values()),
    )
    if summary_header_row < 0:
        return {}

    summary_headers = row_labels(sheet, summary_header_row)
    project_col = find_column(summary_headers, "Наименование проекта")
    plan_col = find_column(summary_headers, "Бюджет проекта без резервов")
    reserve_title_row = find_row(
        sheet,
        lambda labels, row: summary_header_row <= row < detail_header_row
        and any("утвержденные резервы" in value for value in labels.values()),
    )
    reserve_level_row = reserve_title_row + 1 if reserve_title_row > 0 else summary_header_row + 1
    reserve_labels = row_labels(sheet, reserve_level_row)
    reserve_title_labels = row_labels(sheet, reserve_title_row) if reserve_title_row > 0 else {}
    first_reserve_col = min(
        (col for col, value in reserve_title_labels.items() if "утвержденные резервы" in value),
        default=1,
    )
    reserve_columns = {level: find_column(reserve_labels, level, first_reserve_col) for level in RESERVE_LEVELS}

    projects: dict[str, ProjectSummary] = {}
    for row in range(summary_header_row + 1, detail_header_row):
        project = text_value(safe_cell(sheet, row, project_col))
        if not project or "итого" in normalize(project):
            continue

        approved_reserves = {
            level: numeric_value(safe_cell(sheet, row, col)) or 0
            for level, col in reserve_columns.items()
            if col > 0
        }
        plan = numeric_value(safe_cell(sheet, row, plan_col)) or 0
        if not plan and not any(approved_reserves.values()):
            continue

        projects[project] = ProjectSummary(project=project, plan=plan, approved_reserves={**{level: 0 for level in RESERVE_LEVELS}, **approved_reserves})

    return projects


def read_openings(sheet: Worksheet, detail_header_row: int, projects: dict[str, ProjectSummary], source_file: str) -> list[OpeningRow]:
    headers = row_labels(sheet, detail_header_row)
    columns = {
        "project": find_column(headers, "Проект"),
        "cfo": find_column(headers, "ЦФО"),
        "level": find_column(headers, "Уровень резерва"),
        "date": find_column(headers, "Дата вскрытия"),
        "reserve": find_column(headers, "Резерв уровня ЦФО по ПД"),
        "amount": find_column(headers, "Сумма вскрытия"),
        "reason": find_column(headers, "Причина"),
        "counterparty": find_column(headers, "Контрагент"),
        "opening_type": find_column(headers, "Тип вскрытия"),
        "contract": find_column(headers, "Договор"),
        "remainder": find_column(headers, "Остаток от резерва"),
        "economist_comment": find_column(headers, "Комментарий экономиста"),
    }

    default_project = next(iter(projects.keys()), "") if len(projects) == 1 else ""
    openings: list[OpeningRow] = []

    for row in range(detail_header_row + 1, sheet.max_row + 1):
        project = text_value(safe_cell(sheet, row, columns["project"])) or default_project
        cfo = text_value(safe_cell(sheet, row, columns["cfo"]))
        level = text_value(safe_cell(sheet, row, columns["level"]))
        amount = numeric_value(safe_cell(sheet, row, columns["amount"]))

        if not project or normalize(project) == "итого" or not cfo or not level or not amount:
            continue

        opening = OpeningRow(
            source_file=source_file,
            project=project,
            cfo=cfo,
            level=level,
            date_year=date_year(safe_cell(sheet, row, columns["date"]).value if columns["date"] > 0 else None),
            reserve_by_pd=numeric_value(safe_cell(sheet, row, columns["reserve"])) or "",
            amount=amount,
            reason=text_value(safe_cell(sheet, row, columns["reason"])),
            counterparty=text_value(safe_cell(sheet, row, columns["counterparty"])),
            opening_type=text_value(safe_cell(sheet, row, columns["opening_type"])),
            contract=text_value(safe_cell(sheet, row, columns["contract"])),
            remainder=numeric_value(safe_cell(sheet, row, columns["remainder"]))
            or text_value(safe_cell(sheet, row, columns["remainder"])),
            economist_comment=text_value(safe_cell(sheet, row, columns["economist_comment"])),
        )
        openings.append(opening)
        if opening.project not in projects:
            projects[opening.project] = ProjectSummary(project=opening.project)

    return openings


def parse_reserve_workbook(content: bytes, source_file: str) -> ParsedFile:
    # data_only=True читает сохранённый Excel результат формул, а не текст формулы.
    workbook = load_workbook(io.BytesIO(content), data_only=True, keep_vba=True)
    formula_workbook = load_workbook(io.BytesIO(content), data_only=False, keep_vba=True)
    sheet = find_sheet(workbook, INPUT_SHEET_NAME)
    if sheet is None:
        raise ValueError(f'В файле "{source_file}" не найден лист "{INPUT_SHEET_NAME}".')
    apply_formula_fallbacks(sheet, find_sheet(formula_workbook, INPUT_SHEET_NAME))

    detail_header_row = find_row(
        sheet,
        lambda labels, _row: " ".join(labels.values()).find("сумма вскрытия") >= 0
        and any(value == "проект" for value in labels.values())
        and any(value == "цфо" for value in labels.values()),
    )
    if detail_header_row < 0:
        raise ValueError(f'В файле "{source_file}" не найдена таблица вскрытий.')

    projects = read_project_summaries(sheet, detail_header_row)
    openings = read_openings(sheet, detail_header_row, projects, source_file)
    return ParsedFile(source_file=source_file, projects=projects, openings=openings)


def clear_or_create_sheet(workbook: Workbook, name: str, index: int | None = None) -> Worksheet:
    if name in workbook.sheetnames:
        old_sheet = workbook[name]
        old_index = workbook.sheetnames.index(name)
        workbook.remove(old_sheet)
        sheet = workbook.create_sheet(name, old_index)
    elif index is not None:
        sheet = workbook.create_sheet(name, index)
    else:
        sheet = workbook.create_sheet(name)
    return sheet


BLK = "FF000000"
SIDE_M = Side(style="medium", color=BLK)
SIDE_T = Side(style="thin", color=BLK)
SIDE_H = Side(style="hair", color=BLK)
NO_BORDER = Border()
FILL_HEADER = PatternFill(patternType="solid", fgColor="FFD9D9D9")
FILL_EMPTY = PatternFill(fill_type=None)

CFO_STYLES = {
    "строители": ("FF9DC3E6", "FF000000"),
    "уэиэ": ("FFC65911", "FF000000"),
    "ссп": ("FF000000", "FFFFFFFF"),
    "срнп": ("FFFF0000", "FF000000"),
    "сааж": ("FFE2F0D9", "FF000000"),
    "сср": ("FFFF0000", "FF000000"),
    "hr": ("FF7030A0", "FF000000"),
    "сбипр": ("FF000000", "FFFFFFFF"),
}

# Лист «Справка по резервам »: пустые строки между блоками
ROWS_BETWEEN_PROJECTS = 1
ROWS_GAP_DETAIL_HEADER_TO_BODY = 0  # после строки «ЦФО» сразу идёт тело группировки

FONT_TNR = "Times New Roman"
FONT_HDR = Font(name=FONT_TNR, size=12, bold=True, color="FF000000")
FONT_DATA = Font(name=FONT_TNR, size=11, bold=False, color="FF000000")
FONT_DATA_BOLD = Font(name=FONT_TNR, size=11, bold=True, color="FF000000")

NUM_FMT_RU = "# ##0,00"
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Фиксированная ширина только для «Комментарий»; остальные — по тексту
COMMENT_COL_WIDTH = 52.0


def style_summary_block_headers(sheet: Worksheet, row: int, start_col: int, end_col: int) -> None:
    """Верхняя строка сводки."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = ALIGN_CENTER
        left = SIDE_M if col == start_col else SIDE_T
        right = SIDE_M if col == end_col else SIDE_T
        cell.border = Border(left=left, right=right, top=SIDE_M, bottom=SIDE_T)


def style_summary_reserve_subheader(sheet: Worksheet, row: int, start_col: int, end_col: int) -> None:
    """Строка «План по ПД с резервами» / РП / …"""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = ALIGN_CENTER
        left = SIDE_M if col == start_col else SIDE_T
        right = SIDE_M if col == end_col else SIDE_T
        cell.border = Border(left=left, right=right, top=SIDE_T, bottom=SIDE_T)


def style_gutter_row(sheet: Worksheet, row: int, start_col: int, end_col: int) -> None:
    """Пустая строка-отступ: без рамок (только вертикальный зазор на листе)."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_EMPTY
        cell.border = NO_BORDER


def style_summary_data_row(
    sheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    *,
    bottom: Side = SIDE_T,
) -> None:
    """Строки данных сводного блока (проект и резервы)."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_EMPTY
        cell.font = FONT_DATA
        left = SIDE_M if col == start_col else SIDE_T
        right = SIDE_M if col == end_col else SIDE_T
        cell.border = Border(left=left, right=right, top=SIDE_T, bottom=bottom)
        if col in (4, 5, 6, 7, 8, 9):
            cell.alignment = ALIGN_CENTER
            cell.number_format = NUM_FMT_RU
        else:
            cell.alignment = ALIGN_CENTER


def style_detail_header_template(sheet: Worksheet, row: int, start_col: int, end_col: int) -> None:
    """Строка заголовка таблицы вскрытий."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = ALIGN_CENTER
        left = SIDE_M if col == start_col else SIDE_H
        right = SIDE_M if col == end_col else SIDE_H
        cell.border = Border(left=left, right=right, top=SIDE_M, bottom=SIDE_M)

# Колонки детализации: B=ЦФО, D=Уровень, E=Вид работ, F=Сумма, G=Тип, H=Комментарий, I=Дата вскрытия
COL_CFO, COL_LVL, COL_WORK, COL_AMT, COL_TYPE, COL_COMM, COL_YEAR = 2, 4, 5, 6, 7, 8, 9


def _ordered_reserve_levels(by_level: dict[str, list]) -> list[str]:
    out: list[str] = []
    for lv in RESERVE_LEVELS:
        if by_level.get(lv):
            out.append(lv)
    extra = sorted(k for k in by_level if k not in RESERVE_LEVELS and by_level[k])
    out.extend(extra)
    return out


def _build_cfo_structure(openings: list[OpeningRow]) -> list[tuple[str, list[tuple[str, list[OpeningRow]]]]]:
    by_cfo: dict[str, list[OpeningRow]] = defaultdict(list)
    for opening in openings:
        by_cfo[opening.cfo].append(opening)
    structure: list[tuple[str, list[tuple[str, list[OpeningRow]]]]] = []
    for cfo in sorted(by_cfo.keys()):
        by_level: dict[str, list[OpeningRow]] = defaultdict(list)
        for opening in by_cfo[cfo]:
            by_level[opening.level].append(opening)
        levels = _ordered_reserve_levels(by_level)
        groups = [(lv, by_level[lv]) for lv in levels]
        structure.append((cfo, groups))
    return structure


def _count_grouped_detail_rows(openings: list[OpeningRow]) -> int:
    if not openings:
        return 1
    total = 0
    for _cfo, groups in _build_cfo_structure(openings):
        total += 1
        for _lv, items in groups:
            total += 1 + len(items)
    return total


def _str_len_for_width(value: object) -> int:
    return len(str(value).replace("\n", " ")) if value is not None else 0


def compute_detail_column_widths(projects: list[ProjectSummary]) -> dict[str, float]:
    """Ширина по самому длинному тексту в столбце; колонка «Комментарий» (H) — фикс."""
    samples: dict[int, list[str]] = {c: [] for c in range(2, 10)}

    top_headers = [
        "Наименование проекта",
        "Описание проекта ",
        "План по ПД",
        "Вскрыто резервов итого",
        "В т.ч. идеологические изменения",
        "В т.ч. удорожание проекта",
        "В т.ч. техническое вскрытие",
        "Вскрыто резервов 2025",
    ]
    for i, t in enumerate(top_headers):
        samples[2 + i].append(t)

    res_top = ["План по ПД с резервами", "РП", "РГП", "ЗГД", "ГД", "Итого резервы"]
    for i, t in enumerate(res_top):
        samples[4 + i].append(t)

    det_headers = ["ЦФО", "", "Уровень резерва", "Вид работ", "Сумма вскрытия", "Тип вскрытия", "Комментарий", "Дата вскрытия"]
    for i, t in enumerate(det_headers):
        samples[2 + i].append(t)

    for project in projects:
        samples[2].append(project.project)
        samples[3].append(project.description or "")
        samples[4].append(str(project.plan))
        for i, lv in enumerate(RESERVE_LEVELS):
            samples[5 + i].append(str(project.approved_reserves.get(lv, 0)))

        for opening in project.openings:
            samples[COL_CFO].append(opening.cfo)
            samples[COL_LVL].append(opening.level)
            samples[COL_WORK].append(opening.work_type or "")
            samples[COL_AMT].append(f"{opening.amount:.2f}")
            samples[COL_TYPE].append(opening.opening_type or "")
            samples[COL_COMM].append(opening.economist_comment or "")
            samples[COL_YEAR].append(str(opening.date_year))

    out: dict[str, float] = {}
    for col in range(2, 10):
        letter = get_column_letter(col)
        if col == COL_COMM:
            out[letter] = COMMENT_COL_WIDTH
            continue
        longest = max((_str_len_for_width(s) for s in samples[col]), default=6)
        w = max(9.0, min(46.0, longest * 1.05 + 2.4))
        out[letter] = round(w, 2)
    return out


def _style_detail_body_row(
    sheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Обычная строка детализации."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_EMPTY
        cell.font = FONT_DATA
        left = SIDE_M if col == start_col else SIDE_H
        right = SIDE_M if col == end_col else SIDE_H
        cell.border = Border(left=left, right=right, top=SIDE_H, bottom=SIDE_H)
        if col == COL_AMT:
            cell.alignment = ALIGN_CENTER
            cell.number_format = NUM_FMT_RU
        else:
            cell.alignment = ALIGN_CENTER


def _style_group_subtotal_row(
    sheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Итог по ЦФО / по уровню: жирный в ключевых колонках."""
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = FILL_EMPTY
        cell.font = FONT_DATA_BOLD if col in (COL_CFO, COL_LVL, COL_AMT) else FONT_DATA
        left = SIDE_M if col == start_col else SIDE_H
        right = SIDE_M if col == end_col else SIDE_H
        cell.border = Border(left=left, right=right, top=SIDE_H, bottom=SIDE_H)
        if col == COL_AMT:
            cell.alignment = ALIGN_CENTER
            cell.number_format = NUM_FMT_RU
        else:
            cell.alignment = ALIGN_CENTER


def _apply_detail_outline(sheet: Worksheet, row: int, level: int) -> None:
    sheet.row_dimensions[row].outline_level = level


def _cfo_colors(cfo: str) -> tuple[str, str] | None:
    return CFO_STYLES.get(normalize(cfo).replace(" ", ""))


def _apply_cfo_color(sheet: Worksheet, row: int, cfo: str, *, whole_row: bool = False) -> None:
    colors = _cfo_colors(cfo)
    if not colors:
        return
    fill_rgb, font_rgb = colors
    start_col, end_col = (2, 9) if whole_row else (2, 3)
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = PatternFill(patternType="solid", fgColor=fill_rgb)
        cell.font = Font(name=FONT_TNR, size=11, bold=True, color=font_rgb)
        cell.alignment = ALIGN_CENTER


def _merge_cfo_cells(sheet: Worksheet, row: int) -> None:
    """В детализации колонка ЦФО визуально занимает B:C."""
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)


def _clear_row_outline(sheet: Worksheet, row: int) -> None:
    """Пустая строка не должна входить в иерархию группировки."""
    sheet.row_dimensions[row].outline_level = 0


def _set_outline_summary_above(sheet: Worksheet) -> None:
    try:
        from openpyxl.worksheet.properties import Outline

        sp = sheet.sheet_properties
        if sp.outlinePr is None:
            sp.outlinePr = Outline(summaryBelow=False, summaryRight=False, applyStyles=False)
        else:
            sp.outlinePr.summaryBelow = False
    except Exception:
        pass


def compute_flat_column_widths(openings: list[OpeningRow]) -> dict[str, float]:
    hdrs = ["Уровень резерва", "Вид работ", "Сумма вскрытия", "Тип вскрытия", "Комментарий", "Дата вскрытия", "ЦФО"]
    samples: dict[int, list[str]] = {c: [] for c in range(1, 8)}
    for i, h in enumerate(hdrs):
        samples[i + 1].append(h)
    for opening in openings:
        samples[1].append(opening.level)
        samples[2].append(opening.work_type or "")
        samples[3].append(f"{opening.amount:.2f}")
        samples[4].append(opening.opening_type or "")
        samples[5].append(opening.economist_comment or "")
        samples[6].append(str(opening.date_year))
        samples[7].append(opening.cfo)
    out: dict[str, float] = {}
    for col in range(1, 8):
        letter = get_column_letter(col)
        if col == 5:
            out[letter] = COMMENT_COL_WIDTH
            continue
        longest = max((_str_len_for_width(s) for s in samples[col]), default=6)
        out[letter] = round(max(9.0, min(46.0, longest * 1.05 + 2.4)), 2)
    return out


def _style_flat_data_row(sheet: Worksheet, row: int) -> None:
    for col in range(1, 8):
        cell = sheet.cell(row=row, column=col)
        cell.font = FONT_DATA
        cell.border = NO_BORDER
        if col == 3:
            cell.number_format = NUM_FMT_RU
            cell.alignment = ALIGN_RIGHT
        elif col == 6:
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT


def _style_flat_total_row(sheet: Worksheet, row: int) -> None:
    for col in range(1, 8):
        cell = sheet.cell(row=row, column=col)
        cell.font = FONT_DATA_BOLD
        cell.border = NO_BORDER
        if col == 3:
            cell.number_format = NUM_FMT_RU
            cell.alignment = ALIGN_RIGHT
        else:
            cell.alignment = ALIGN_LEFT


def set_widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_reserve_sheet(workbook: Workbook, projects: list[ProjectSummary]) -> dict[str, int]:
    sheet = clear_or_create_sheet(workbook, TARGET_SHEET_NAME, 0)
    _set_outline_summary_above(sheet)
    summary_rows: dict[str, int] = {}
    row = 3

    for project in projects:
        if row > 3:
            row += ROWS_BETWEEN_PROJECTS
            base = row - ROWS_BETWEEN_PROJECTS
            for i in range(ROWS_BETWEEN_PROJECTS):
                gr = base + i
                style_gutter_row(sheet, gr, 2, 9)
                _clear_row_outline(sheet, gr)

        structure = _build_cfo_structure(project.openings)
        n_detail = _count_grouped_detail_rows(project.openings)
        summary_row = row + 1
        reserves_row = row + 3
        detail_header_row = row + 5
        detail_first = detail_header_row + 1 + ROWS_GAP_DETAIL_HEADER_TO_BODY
        detail_last = detail_first + n_detail - 1
        summary_rows[project.project] = summary_row

        headers = [
            "Наименование проекта",
            "Описание проекта ",
            "План по ПД",
            "Вскрыто резервов итого",
            "В т.ч. идеологические изменения",
            "В т.ч. удорожание проекта",
            "В т.ч. техническое вскрытие",
            "Вскрыто резервов 2025",
        ]
        for offset, value in enumerate(headers, start=2):
            sheet.cell(row=row, column=offset, value=value)
        style_summary_block_headers(sheet, row, 2, 9)

        sheet.cell(row=summary_row, column=2, value=project.project)
        sheet.cell(row=summary_row, column=3, value=project.description)
        sheet.cell(row=summary_row, column=4, value=project.plan)
        sheet.cell(row=summary_row, column=5, value=f"=SUM(F{summary_row}:H{summary_row})")
        sheet.cell(
            row=summary_row,
            column=6,
            value=f'=SUMIF(G{detail_first}:G{detail_last},"Идеологическое изменение",F{detail_first}:F{detail_last})',
        )
        sheet.cell(
            row=summary_row,
            column=7,
            value=f'=SUMIF(G{detail_first}:G{detail_last},"Удорожание",F{detail_first}:F{detail_last})',
        )
        sheet.cell(
            row=summary_row,
            column=8,
            value=f'=SUMIF(G{detail_first}:G{detail_last},"Техническое",F{detail_first}:F{detail_last})',
        )
        sheet.cell(
            row=summary_row,
            column=9,
            value=f"=SUMIF(I{detail_first}:I{detail_last},2025,F{detail_first}:F{detail_last})",
        )
        style_summary_data_row(sheet, summary_row, 2, 9)

        reserve_headers = ["План по ПД с резервами", "РП", "РГП", "ЗГД", "ГД", "Итого резервы"]
        for offset, value in enumerate(reserve_headers, start=4):
            sheet.cell(row=row + 2, column=offset, value=value)
        style_summary_reserve_subheader(sheet, row + 2, 2, 9)

        sheet.cell(row=reserves_row, column=4, value=f"=D{summary_row}+I{reserves_row}")
        for index, level in enumerate(RESERVE_LEVELS, start=5):
            sheet.cell(row=reserves_row, column=index, value=project.approved_reserves.get(level, 0))
        sheet.cell(row=reserves_row, column=9, value=f"=SUM(E{reserves_row}:H{reserves_row})")
        style_summary_data_row(sheet, reserves_row, 2, 9, bottom=SIDE_M)

        # 1 пустая строка между верхним блоком (резервы) и строкой заголовков таблицы вскрытий
        style_gutter_row(sheet, row + 4, 2, 9)
        _clear_row_outline(sheet, row + 4)

        detail_headers = ["ЦФО", "", "Уровень резерва", "Вид работ", "Сумма вскрытия", "Тип вскрытия", "Комментарий", "Дата вскрытия"]
        for offset, value in enumerate(detail_headers, start=2):
            sheet.cell(row=detail_header_row, column=offset, value=value)
        style_detail_header_template(sheet, detail_header_row, 2, 9)
        _merge_cfo_cells(sheet, detail_header_row)

        for k in range(ROWS_GAP_DETAIL_HEADER_TO_BODY):
            gap_r = detail_header_row + 1 + k
            style_gutter_row(sheet, gap_r, 2, 9)
            _clear_row_outline(sheet, gap_r)

        r = detail_first
        if not project.openings:
            sheet.cell(row=r, column=COL_CFO, value="(нет строк вскрытия)")
            sheet.cell(row=r, column=COL_AMT, value=0)
            _style_detail_body_row(sheet, r, 2, 9)
            _merge_cfo_cells(sheet, r)
            _apply_detail_outline(sheet, r, 0)
            r += 1
        else:
            for cfo, groups in structure:
                cfo_row = r
                r += 1
                subtotal_row_refs: list[int] = []

                for level_name, items in groups:
                    sub_row = r
                    r += 1
                    first_detail = r
                    for opening in items:
                        sheet.cell(row=r, column=COL_CFO, value=opening.cfo)
                        sheet.cell(row=r, column=COL_LVL, value=opening.level)
                        sheet.cell(row=r, column=COL_WORK, value=opening.work_type)
                        sheet.cell(row=r, column=COL_AMT, value=opening.amount)
                        sheet.cell(row=r, column=COL_TYPE, value=opening.opening_type)
                        sheet.cell(row=r, column=COL_COMM, value=opening.economist_comment)
                        sheet.cell(row=r, column=COL_YEAR, value=opening.date_year)
                        _style_detail_body_row(sheet, r, 2, 9)
                        _apply_cfo_color(sheet, r, opening.cfo)
                        _merge_cfo_cells(sheet, r)
                        _apply_detail_outline(sheet, r, 2)
                        r += 1
                    last_detail = r - 1

                    sheet.cell(row=sub_row, column=COL_CFO, value=cfo)
                    sheet.cell(row=sub_row, column=COL_LVL, value=level_name)
                    if first_detail <= last_detail:
                        sheet.cell(row=sub_row, column=COL_AMT, value=f"=SUM(F{first_detail}:F{last_detail})")
                    else:
                        sheet.cell(row=sub_row, column=COL_AMT, value=0)
                    _style_group_subtotal_row(sheet, sub_row, 2, 9)
                    _apply_cfo_color(sheet, sub_row, cfo)
                    _merge_cfo_cells(sheet, sub_row)
                    _apply_detail_outline(sheet, sub_row, 1)
                    subtotal_row_refs.append(sub_row)

                cfo_formula = "=" + "+".join(f"F{x}" for x in subtotal_row_refs) if subtotal_row_refs else "=0"
                sheet.cell(row=cfo_row, column=COL_CFO, value=cfo)
                sheet.cell(row=cfo_row, column=COL_AMT, value=cfo_formula)
                _style_group_subtotal_row(sheet, cfo_row, 2, 9)
                _apply_cfo_color(sheet, cfo_row, cfo, whole_row=True)
                _merge_cfo_cells(sheet, cfo_row)
                _apply_detail_outline(sheet, cfo_row, 0)

        last_written = r - 1
        if last_written != detail_last:
            raise RuntimeError(
                f"Несовпадение числа строк блока проекта {project.project}: ожидалось {detail_last}, запись до {last_written}"
            )

        row = detail_last + 1

    set_widths(sheet, compute_detail_column_widths(projects))
    return summary_rows


def write_flat_sheet(workbook: Workbook, openings: list[OpeningRow]) -> None:
    sheet = clear_or_create_sheet(workbook, FLAT_SHEET_NAME)
    headers = ["Уровень резерва", "Вид работ", "Сумма вскрытия", "Тип вскрытия", "Комментарий", "Дата вскрытия", "ЦФО"]
    sheet.append(headers)
    style_detail_header_template(sheet, 1, 1, len(headers))
    first_data = 2
    for opening in openings:
        sheet.append(
            [
                opening.level,
                opening.work_type,
                opening.amount,
                opening.opening_type,
                opening.economist_comment,
                opening.date_year,
                opening.cfo,
            ]
        )
        _style_flat_data_row(sheet, sheet.max_row)
    if openings:
        last_data = sheet.max_row
        total_row = last_data + 1
        sheet.cell(row=total_row, column=1, value="ИТОГО")
        sheet.cell(row=total_row, column=3, value=f"=SUBTOTAL(9,C{first_data}:C{last_data})")
        _style_flat_total_row(sheet, total_row)
    set_widths(sheet, compute_flat_column_widths(openings))


def write_summary_sheet(workbook: Workbook, projects: list[ProjectSummary], summary_rows: dict[str, int]) -> None:
    sheet = clear_or_create_sheet(workbook, SUMMARY_SHEET_NAME)
    sheet.cell(row=1, column=2, value="добавить описание проекта + техническое вскрытие")
    headers = [
        "Наименование проекта",
        "Описание проекта",
        "План по ПД",
        "Вскрыто резервов итого",
        "В т.ч. идеологические изменения",
        "В т.ч. удорожание проекта",
        "В т.ч. техническое вскрытие",
        "Вскрыто резервов 2025",
    ]
    for offset, value in enumerate(headers, start=2):
        sheet.cell(row=2, column=offset, value=value)
    style_summary_block_headers(sheet, 2, 2, 9)

    last_project_row = len(projects) + 3
    sheet.cell(row=3, column=2, value="ИТОГО:")
    for col in range(4, 10):
        letter = get_column_letter(col)
        sheet.cell(row=3, column=col, value=f"=SUBTOTAL(9,{letter}4:{letter}{last_project_row})")

    for index, project in enumerate(projects, start=4):
        source_row = summary_rows[project.project]
        sheet.cell(row=index, column=2, value=project.project)
        for col in range(3, 10):
            letter = get_column_letter(col)
            sheet.cell(row=index, column=col, value=f"='{TARGET_SHEET_NAME}'!{letter}{source_row}")
    set_widths(sheet, compute_detail_column_widths(projects))


def merge_parsed_files(parsed_files: list[ParsedFile]) -> list[ProjectSummary]:
    projects: dict[str, ProjectSummary] = {}
    for parsed in parsed_files:
        for project_name, project in parsed.projects.items():
            if project_name not in projects:
                projects[project_name] = ProjectSummary(
                    project=project.project,
                    plan=project.plan,
                    description=project.description,
                    approved_reserves=copy(project.approved_reserves),
                )
            else:
                existing = projects[project_name]
                existing.plan = existing.plan or project.plan
                existing.description = existing.description or project.description
                for level in RESERVE_LEVELS:
                    existing.approved_reserves[level] = existing.approved_reserves.get(level, 0) or project.approved_reserves.get(level, 0)

        for opening in parsed.openings:
            projects[opening.project].openings.append(opening)

    result = sorted(projects.values(), key=lambda item: item.project)
    for project in result:
        project.openings.sort(key=lambda item: (item.cfo, item.level, str(item.date_year)))
    return result


def build_consolidated_workbook(
    parsed_files: list[ParsedFile],
) -> tuple[bytes, list[ProjectSummary], list[OpeningRow]]:
    """Собирает новую книгу `.xlsx` только из загруженных файлов вскрытия резервов."""
    projects = merge_parsed_files(parsed_files)
    openings = [opening for project in projects for opening in project.openings]

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_rows = write_reserve_sheet(workbook, projects)
    write_flat_sheet(workbook, openings)
    write_summary_sheet(workbook, projects, summary_rows)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), projects, openings
